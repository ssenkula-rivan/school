"""
Utility functions for generating reports and exports
"""

from io import BytesIO
from django.http import HttpResponse
from django.template.loader import render_to_string
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


def generate_pdf_report(template_name, context, filename):
    """
    Generate PDF from HTML template using ReportLab
    """
    try:
        from reportlab.lib.pagesizes import letter, A4
        from reportlab.lib import colors
        from reportlab.lib.units import inch
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        elements = []
        styles = getSampleStyleSheet()
        
        # Add title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#667eea'),
            spaceAfter=30,
            alignment=1  # Center
        )
        
        if 'title' in context:
            title = Paragraph(context['title'], title_style)
            elements.append(title)
            elements.append(Spacer(1, 0.2*inch))
        
        # Return PDF response
        doc.build(elements)
        buffer.seek(0)
        
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
        
    except ImportError:
        # Fallback if reportlab not installed
        return HttpResponse("PDF generation library not installed", status=500)


def generate_excel_report(data, headers, filename, sheet_name='Report'):
    """
    Generate Excel file from data
    
    Args:
        data: List of lists/tuples containing row data
        headers: List of column headers
        filename: Output filename
        sheet_name: Excel sheet name
    """
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_name
    
    # Style definitions
    header_font = Font(bold=True, color='FFFFFF', size=12)
    header_fill = PatternFill(start_color='667eea', end_color='667eea', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Write data
    for row_num, row_data in enumerate(data, 2):
        for col_num, cell_value in enumerate(row_data, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value
            cell.border = border
            cell.alignment = Alignment(horizontal='left', vertical='center')
    
    # Auto-adjust column widths
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        worksheet.column_dimensions[column_letter].width = adjusted_width
    
    # Save to BytesIO
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    
    response = HttpResponse(
        buffer,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response


def export_report_card_pdf(report_card):
    """Generate PDF for student report card"""
    from academics.models import Mark
    
    # Get all marks for this report card
    marks = Mark.objects.filter(
        student=report_card.student,
        exam__academic_year=report_card.academic_year,
        exam__term=report_card.term
    ).select_related('subject', 'exam')
    
    context = {
        'title': f'Report Card - {report_card.student.get_full_name()}',
        'report_card': report_card,
        'marks': marks,
        'school': report_card.student.grade,
        'generated_date': datetime.now()
    }
    
    filename = f'report_card_{report_card.student.admission_number}_{report_card.term}.pdf'
    return generate_pdf_report('reports/report_card_pdf.html', context, filename)


def export_fee_receipt_pdf(payment):
    """Generate PDF for fee payment receipt"""
    context = {
        'title': 'Fee Payment Receipt',
        'payment': payment,
        'student': payment.student,
        'generated_date': datetime.now()
    }
    
    filename = f'receipt_{payment.receipt_number}.pdf'
    return generate_pdf_report('reports/fee_receipt_pdf.html', context, filename)


def export_students_excel(students):
    """Export students list to Excel"""
    headers = [
        'Admission Number', 'Full Name', 'Grade', 'Gender', 
        'Date of Birth', 'Guardian Name', 'Guardian Phone', 
        'Status', 'Admission Date'
    ]
    
    data = []
    for student in students:
        data.append([
            student.admission_number,
            student.get_full_name(),
            str(student.grade),
            student.get_gender_display(),
            student.date_of_birth.strftime('%Y-%m-%d'),
            student.guardian_name,
            student.guardian_phone,
            student.get_status_display(),
            student.admission_date.strftime('%Y-%m-%d')
        ])
    
    filename = f'students_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    return generate_excel_report(data, headers, filename, 'Students')


def export_attendance_excel(attendance_records):
    """Export attendance records to Excel"""
    headers = [
        'Student', 'Admission Number', 'Date', 'Status', 
        'Marked By', 'Remarks'
    ]
    
    data = []
    for record in attendance_records:
        data.append([
            record.student.get_full_name(),
            record.student.admission_number,
            record.date.strftime('%Y-%m-%d'),
            record.get_status_display(),
            record.marked_by.get_full_name() if record.marked_by else 'N/A',
            record.remarks
        ])
    
    filename = f'attendance_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    return generate_excel_report(data, headers, filename, 'Attendance')


def export_financial_summary_excel(payments, start_date, end_date):
    """Export financial summary to Excel"""
    headers = [
        'Receipt Number', 'Student', 'Amount', 'Payment Method',
        'Payment Date', 'Status', 'Received By'
    ]
    
    data = []
    total_amount = 0
    
    for payment in payments:
        data.append([
            payment.receipt_number,
            payment.student.get_full_name(),
            float(payment.amount_paid),
            payment.get_payment_method_display(),
            payment.payment_date.strftime('%Y-%m-%d'),
            payment.get_payment_status_display(),
            payment.received_by.get_full_name() if payment.received_by else 'N/A'
        ])
        total_amount += payment.amount_paid
    
    # Add summary row
    data.append([])
    data.append(['TOTAL', '', float(total_amount), '', '', '', ''])
    
    filename = f'financial_summary_{start_date}_{end_date}.xlsx'
    return generate_excel_report(data, headers, filename, 'Financial Summary')


def generate_academic_analytics_data(academic_year, term):
    """Generate data for academic analytics dashboard"""
    from academics.models import Mark, ReportCard
    from core.models import Student
    from django.db.models import Avg, Count, Q
    
    # Overall statistics
    total_students = Student.objects.filter(status='active').count()
    
    # Grade distribution
    grade_distribution = Mark.objects.filter(
        exam__academic_year=academic_year,
        exam__term=term
    ).values('grade').annotate(count=Count('id')).order_by('grade')
    
    # Average performance by subject
    subject_performance = Mark.objects.filter(
        exam__academic_year=academic_year,
        exam__term=term
    ).values('subject__name').annotate(
        avg_percentage=Avg('percentage'),
        avg_grade_point=Avg('grade_point')
    ).order_by('-avg_percentage')
    
    # Top performers
    top_performers = ReportCard.objects.filter(
        academic_year=academic_year,
        term=term
    ).order_by('-average_percentage')[:10]
    
    # Attendance statistics
    from academics.models import StudentAttendance
    attendance_stats = StudentAttendance.objects.filter(
        date__gte=academic_year.start_date,
        date__lte=academic_year.end_date
    ).values('status').annotate(count=Count('id'))
    
    return {
        'total_students': total_students,
        'grade_distribution': list(grade_distribution),
        'subject_performance': list(subject_performance),
        'top_performers': top_performers,
        'attendance_stats': list(attendance_stats)
    }
