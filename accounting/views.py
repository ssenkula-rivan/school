from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse, JsonResponse
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db.models import Sum, Count, Q, Prefetch
from django.utils.timezone import datetime
from django.utils.dateparse import parse_date
from django.template.loader import get_template
from django.views.decorators.csrf import csrf_exempt
from django.conf import settings
from django.db import models
from datetime import date, timedelta
from django.db.models import Sum, Count, Q
import io
import json
import time
import os
import openpyxl
from openpyxl.utils import get_column_letter

# Optional imports for PDF generation
try:
    from xhtml2pdf import pisa
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False

# Optional import for OpenAI
try:
    from openai import OpenAI
    OPENAI_AVAILABLE = True
except ImportError:
    OPENAI_AVAILABLE = False

from .models import Account, Transaction, JournalEntry, Invoice, Receipt
from .forms import TransactionForm, JournalEntryFormSet, InvoiceForm, ReceiptForm






@login_required

def dashboard(request):
    transactions = Transaction.objects.all().order_by('-date')[:5]

    for tx in transactions:
        total_amount = tx.entries.aggregate(total=Sum('amount'))['total'] or 0
        tx.amount = total_amount

    return render(request, 'accounting/dashboard.html', {'transactions': transactions})


def balance_sheet(request):
    assets = Account.objects.filter(account_type__name='Asset')
    liabilities = Account.objects.filter(account_type__name='Liability')
    equity = Account.objects.filter(account_type__name='Equity')

    total_assets = assets.aggregate(total=models.Sum('balance'))['total'] or 0
    total_liabilities = liabilities.aggregate(total=models.Sum('balance'))['total'] or 0
    total_equity = equity.aggregate(total=models.Sum('balance'))['total'] or 0

    return render(request, 'accounting/balance_sheet.html', {
        'assets': assets,
        'liabilities': liabilities,
        'equity': equity,
        'total_assets': total_assets,
        'total_liabilities': total_liabilities,
        'total_equity': total_equity
    })


def chart_of_accounts(request):
    accounts = Account.objects.all()
    return render(request, 'accounting/chart_of_accounts.html', {
        'accounts': accounts
    })


def transaction_list(request):
    if request.method == 'POST':
        form = TransactionForm(request.POST)
        transaction = form.save(commit=False)
        transaction.user = request.user
        formset = JournalEntryFormSet(request.POST, instance=transaction)

        if form.is_valid() and formset.is_valid():
            total_debits = 0
            total_credits = 0
            valid_entries = []

            for f in formset:
                if not f.cleaned_data or f.cleaned_data.get('DELETE'):
                    continue

                account = f.cleaned_data.get('account')
                amount = f.cleaned_data.get('amount')
                entry_type = f.cleaned_data.get('entry_type')

                if account and amount and entry_type:
                    valid_entries.append(f)
                    if entry_type == 'DEBIT':
                        total_debits += amount
                    elif entry_type == 'CREDIT':
                        total_credits += amount

            if not valid_entries:
                messages.error(request, "You must enter at least one valid journal entry.")
            else:
                transaction.save()

                # Save all user-entered entries
                for f in valid_entries:
                    f.instance.transaction = transaction
                    f.save()

                # Auto-balance if needed
                difference = round(total_debits - total_credits, 2)
                if difference != 0:
                    from .models import Account
                    suspense = Account.objects.get(name='Suspense', user=request.user)

                    entry_type = 'CREDIT' if difference > 0 else 'DEBIT'

                    JournalEntry.objects.create(
                        transaction=transaction,
                        account=suspense,
                        amount=abs(difference),
                        entry_type=entry_type
                    )

                    messages.warning(request, f"Auto-balanced with suspense account: {entry_type} ${abs(difference):.2f}")

                messages.success(request, "Transaction saved successfully.")
                return redirect('dashboard')
    else:
        form = TransactionForm()
        formset = JournalEntryFormSet()

    return render(request, 'accounting/transaction_list.html', {
        'form': form,
        'formset': formset
    })


def transaction_history(request):
    transactions = Transaction.objects.prefetch_related(
        Prefetch('entries', queryset=JournalEntry.objects.select_related('account'))
    ).order_by('-date')

    return render(request, 'accounting/transaction_history.html', {
        'transactions': transactions
    })

from django.utils.dateparse import parse_date

def profit_and_loss(request):
    from .models import JournalEntry, AccountType

    start_date = parse_date(request.GET.get('start_date') or '')
    end_date = parse_date(request.GET.get('end_date') or '')

    # Get or create account types
    revenue_type, _ = AccountType.objects.get_or_create(name='Revenue', defaults={'normal_balance': 'CREDIT'})
    expense_type, _ = AccountType.objects.get_or_create(name='Expense', defaults={'normal_balance': 'DEBIT'})

    revenue_entries = JournalEntry.objects.filter(account__account_type=revenue_type)
    expense_entries = JournalEntry.objects.filter(account__account_type=expense_type)

    if start_date and end_date:
        revenue_entries = revenue_entries.filter(transaction__date__range=(start_date, end_date))
        expense_entries = expense_entries.filter(transaction__date__range=(start_date, end_date))

    revenue_summary = revenue_entries.values('account__name').annotate(total=Sum('amount'))
    expense_summary = expense_entries.values('account__name').annotate(total=Sum('amount'))

    total_revenue = revenue_entries.aggregate(total=Sum('amount'))['total'] or 0
    total_expenses = expense_entries.aggregate(total=Sum('amount'))['total'] or 0
    net_income = total_revenue - total_expenses

    return render(request, 'accounting/profit_and_loss.html', {
        'revenue_summary': revenue_summary,
        'expense_summary': expense_summary,
        'total_revenue': total_revenue,
        'total_expenses': total_expenses,
        'net_income': net_income,
        'start_date': start_date,
        'end_date': end_date,
    })

def export_profit_loss_pdf(request):
    from .models import JournalEntry, AccountType

    start_date = parse_date(request.GET.get('start_date') or '')
    end_date = parse_date(request.GET.get('end_date') or '')

    # Get or create account types
    revenue_type, _ = AccountType.objects.get_or_create(name='Revenue', defaults={'normal_balance': 'CREDIT'})
    expense_type, _ = AccountType.objects.get_or_create(name='Expense', defaults={'normal_balance': 'DEBIT'})

    revenue_entries = JournalEntry.objects.filter(account__account_type=revenue_type)
    expense_entries = JournalEntry.objects.filter(account__account_type=expense_type)

    if start_date and end_date:
        revenue_entries = revenue_entries.filter(transaction__date__range=(start_date, end_date))
        expense_entries = expense_entries.filter(transaction__date__range=(start_date, end_date))

    revenue_summary = revenue_entries.values('account__name').annotate(total=Sum('amount'))
    expense_summary = expense_entries.values('account__name').annotate(total=Sum('amount'))

    total_revenue = revenue_entries.aggregate(total=Sum('amount'))['total'] or 0
    total_expenses = expense_entries.aggregate(total=Sum('amount'))['total'] or 0
    net_income = total_revenue - total_expenses

    if not PDF_AVAILABLE:
        messages.error(request, 'PDF generation is not available. Please install required dependencies.')
        return redirect('profit_and_loss')

    template = get_template('accounting/profit_and_loss_pdf.html')
    html = template.render({
        'revenue_summary': revenue_summary,
        'expense_summary': expense_summary,
        'total_revenue': total_revenue,
        'total_expenses': total_expenses,
        'net_income': net_income,
        'start_date': start_date,
        'end_date': end_date,
    })

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="profit_and_loss.pdf"'
    pisa.CreatePDF(io.StringIO(html), dest=response)

    return response

def export_profit_loss_excel(request):
    from .models import JournalEntry, AccountType

    start_date = parse_date(request.GET.get('start_date') or '')
    end_date = parse_date(request.GET.get('end_date') or '')

    # Get or create account types
    revenue_type, _ = AccountType.objects.get_or_create(name='Revenue', defaults={'normal_balance': 'CREDIT'})
    expense_type, _ = AccountType.objects.get_or_create(name='Expense', defaults={'normal_balance': 'DEBIT'})

    revenue_entries = JournalEntry.objects.filter(account__account_type=revenue_type)
    expense_entries = JournalEntry.objects.filter(account__account_type=expense_type)

    if start_date and end_date:
        revenue_entries = revenue_entries.filter(transaction__date__range=(start_date, end_date))
        expense_entries = expense_entries.filter(transaction__date__range=(start_date, end_date))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Profit and Loss"

    ws.append(["Profit & Loss Report"])
    ws.append(["Period:", f"{start_date} to {end_date}"])
    ws.append([])

    # Revenue Section
    ws.append(["Revenue"])
    revenue_summary = revenue_entries.values('account__name').annotate(total=Sum('amount'))
    for item in revenue_summary:
        ws.append([item['account__name'], float(item['total'])])
    total_revenue = revenue_entries.aggregate(total=Sum('amount'))['total'] or 0
    ws.append(["Total Revenue", float(total_revenue)])
    ws.append([])

    # Expense Section
    ws.append(["Expenses"])
    expense_summary = expense_entries.values('account__name').annotate(total=Sum('amount'))
    for item in expense_summary:
        ws.append([item['account__name'], float(item['total'])])
    total_expenses = expense_entries.aggregate(total=Sum('amount'))['total'] or 0
    ws.append(["Total Expenses", float(total_expenses)])
    ws.append([])

    net_income = total_revenue - total_expenses
    ws.append(["Net Income", float(net_income)])

    # Format column widths
    for col in range(1, 3):
        ws.column_dimensions[get_column_letter(col)].width = 30

    # Send file as response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = "profit_and_loss.xlsx"
    response['Content-Disposition'] = f'attachment; filename={filename}'
    wb.save(response)
    return response

def view_account_transactions(request, account_id):
    account = get_object_or_404(Account, id=account_id)
    start_date = parse_date(request.GET.get('start_date') or '')
    end_date = parse_date(request.GET.get('end_date') or '')

    entries = JournalEntry.objects.filter(account=account)

    if start_date and end_date:
        entries = entries.filter(transaction__date__range=(start_date, end_date))

    total_debits = entries.filter(entry_type='DEBIT').aggregate(total=Sum('amount'))['total'] or 0
    total_credits = entries.filter(entry_type='CREDIT').aggregate(total=Sum('amount'))['total'] or 0
    transaction_count = entries.aggregate(count=Count('transaction', distinct=True))['count'] or 0
    net_movement = total_debits - total_credits

    return render(request, 'accounting/transaction_detail.html', {
        'account': account,
        'entries': entries,
        'start_date': start_date,
        'end_date': end_date,
        'total_debits': total_debits,
        'total_credits': total_credits,
        'transaction_count': transaction_count,
        'net_movement': net_movement,
    })

def quarterly_tax_doc(request, quarter):
    from .models import JournalEntry, AccountType

    quarters = {
        'q1': ('2025-01-01', '2025-03-31'),
        'q2': ('2025-04-01', '2025-06-30'),
        'q3': ('2025-07-01', '2025-09-30'),
        'q4': ('2025-10-01', '2025-12-31'),
    }

    start, end = quarters.get(quarter, ('2025-01-01', '2025-03-31'))
    start_date = parse_date(start)
    end_date = parse_date(end)

    revenue_type = AccountType.objects.get(name='Revenue')
    expense_type = AccountType.objects.get(name='Expense')

    revenue_entries = JournalEntry.objects.filter(
        account__account_type=revenue_type,
        transaction__date__range=(start_date, end_date)
    )

    expense_entries = JournalEntry.objects.filter(
        account__account_type=expense_type,
        transaction__date__range=(start_date, end_date)
    )

    total_revenue = revenue_entries.aggregate(total=Sum('amount'))['total'] or 0
    total_expenses = expense_entries.aggregate(total=Sum('amount'))['total'] or 0
    net_income = total_revenue - total_expenses

    return render(request, 'accounting/tax_doc.html', {
        'quarter': quarter.upper(),
        'start_date': start_date,
        'end_date': end_date,
        'total_revenue': total_revenue,
        'total_expenses': total_expenses,
        'net_income': net_income
    })

def invoice_list(request):
    today = date.today()
    all_invoices = Invoice.objects.all()

    current_invoices = all_invoices.filter(due_date__gte=today)
    past_due_invoices = all_invoices.filter(due_date__lt=today, status='UNPAID')

    # Pie chart summary by status
    status_summary = (
        all_invoices.values('status')
        .annotate(total=models.Sum('amount'))
        .order_by()
    )

    return render(request, 'accounting/invoices.html', {
        'current_invoices': current_invoices,
        'past_due_invoices': past_due_invoices,
        'status_summary': status_summary,
    })
    
def transaction_detail(request, transaction_id):

    transaction = get_object_or_404(Transaction, id=transaction_id)
    entries = transaction.entries.select_related('account')
    return render(request, 'accounting/transaction_detail.html', {
        'transaction': transaction,
        'entries': entries
    })

def add_invoice(request):
    if request.method == 'POST':
        form = InvoiceForm(request.POST, request.FILES)
        if form.is_valid():
            invoice = form.save(commit=False)
            invoice.user = request.user
            invoice.save()
            return redirect('invoice_list')
    else:
        form = InvoiceForm()

    return render(request, 'accounting/add_invoice.html', {'form': form})

def invoice_detail(request, invoice_id):
    invoice = get_object_or_404(Invoice, id=invoice_id)
    return render(request, 'accounting/invoice_detail.html', {'invoice': invoice})


# Initialize OpenAI client only if available
if OPENAI_AVAILABLE:
    try:
        client = OpenAI(api_key=settings.OPENAI_API_KEY)
    except:
        client = None
else:
    client = None


@csrf_exempt
def chat_with_gpt(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            user_message = data.get('message', '')

            # Process GPT reply with delay (for spinner visibility)
            import time
            time.sleep(1.2)  # Processing delay

            reply = f"Processing your request: '{user_message}'"
            return JsonResponse({'reply': reply})

            # ===== Real GPT logic (disabled for now) =====
            # openai.api_key = settings.OPENAI_API_KEY
            # response = openai.ChatCompletion.create(
            #     model="gpt-3.5-turbo",
            #     messages=[
            #         {"role": "system", "content": "You are an assistant for accounting software."},
            #         {"role": "user", "content": user_message},
            #     ]
            # )
            # real_reply = response.choices[0].message.content.strip()
            # return JsonResponse({'reply': real_reply})

        except Exception as e:
            return JsonResponse({'reply': f"Error: {str(e)}"}, status=500)

    return JsonResponse({'reply': 'Only POST method allowed'}, status=400)


def chat_bot_view(request):
    """Render the chatbot UI page."""
    return render(request, 'accounting/chat_bot.html')


# Receipt Management Views
@login_required
def receipt_list(request):
    """Display all receipts"""
    receipts = Receipt.objects.filter(user=request.user)
    
    # Filter by receipt type if provided
    receipt_type = request.GET.get('receipt_type')
    if receipt_type:
        receipts = receipts.filter(receipt_type=receipt_type)
    
    # Filter by category if provided
    category = request.GET.get('category')
    if category:
        receipts = receipts.filter(category=category)
    
    # Filter by date range if provided
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    if start_date:
        receipts = receipts.filter(date__gte=start_date)
    if end_date:
        receipts = receipts.filter(date__lte=end_date)
    
    # Calculate totals by category
    category_totals = Receipt.objects.filter(user=request.user).values('category').annotate(
        total=Sum('amount')
    ).order_by('-total')
    
    total_amount = receipts.aggregate(total=Sum('amount'))['total'] or 0
    
    context = {
        'receipts': receipts,
        'category_totals': category_totals,
        'total_amount': total_amount,
        'selected_receipt_type': receipt_type,
        'selected_category': category,
        'start_date': start_date,
        'end_date': end_date,
    }
    return render(request, 'accounting/receipt_list.html', context)

@login_required
def add_receipt(request):
    """Add a new receipt"""
    if request.method == 'POST':
        form = ReceiptForm(request.POST, request.FILES)
        if form.is_valid():
            receipt = form.save(commit=False)
            receipt.user = request.user
            receipt.save()
            messages.success(request, 'Receipt added successfully!')
            return redirect('receipt_list')
    else:
        form = ReceiptForm()
    
    return render(request, 'accounting/add_receipt.html', {'form': form})

@login_required
def receipt_detail(request, receipt_id):
    """View receipt details"""
    receipt = get_object_or_404(Receipt, id=receipt_id, user=request.user)
    return render(request, 'accounting/receipt_detail.html', {'receipt': receipt})

@login_required
def edit_receipt(request, receipt_id):
    """Edit an existing receipt"""
    receipt = get_object_or_404(Receipt, id=receipt_id, user=request.user)
    
    if request.method == 'POST':
        form = ReceiptForm(request.POST, request.FILES, instance=receipt)
        if form.is_valid():
            form.save()
            messages.success(request, 'Receipt updated successfully!')
            return redirect('receipt_detail', receipt_id=receipt.id)
    else:
        form = ReceiptForm(instance=receipt)
    
    return render(request, 'accounting/edit_receipt.html', {'form': form, 'receipt': receipt})

@login_required
def delete_receipt(request, receipt_id):
    """Delete a receipt"""
    receipt = get_object_or_404(Receipt, id=receipt_id, user=request.user)
    
    if request.method == 'POST':
        receipt.delete()
        messages.success(request, 'Receipt deleted successfully!')
        return redirect('receipt_list')
    
    return render(request, 'accounting/delete_receipt.html', {'receipt': receipt})
